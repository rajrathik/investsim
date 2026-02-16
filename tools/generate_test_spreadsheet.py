"""
Generate a multi-sheet Excel workbook that mirrors the Portfolio Simulator
calculation logic step-by-step.  Users type tickers + allocations; the script
pulls real historical data from the SQL Server DB and builds five sheets:

  1. Setup        - parameters + raw price/dividend/MM data from DB
  2. Monthly Sim  - full month-by-month simulation with Excel formulas
  3. Year Summary - annual aggregations (SUMIFS from Monthly Sim)
  4. Tax Impact   - yearly tax on dividends and MM interest
  5. Annual Return- pre-tax return with DCA mid-year approx

Key behaviors:
  - Round-lot (integer) share buying: INT(accumulated / high_price)
  - Per-ticker accumulate-then-buy: each ticker keeps its own unspent $
    in a bucket until enough to buy a whole share (no aggregate carryover)
  - Simulation period: January of (current_year - N) through December of
    prior complete year (never includes the current partial year)

Run:
    cd C:\\Raj\\python\\portfolio-simulator
    venv\\Scripts\\python tools\\generate_test_spreadsheet.py
"""

import sys, os, math
from datetime import date
from collections import defaultdict

# --- path setup so we can import from backend/app/ ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, os.path.join(PROJECT_ROOT, "backend"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from app.database import SessionLocal
from app.models import Ticker, MonthlyPrice, Dividend
from app.mm_rates import MonthlyMoneyMarketRate

# ============================================================
# STYLE CONSTANTS
# ============================================================
DARK_BG   = PatternFill("solid", fgColor="1E293B")
ACCENT_BG = PatternFill("solid", fgColor="10B981")
GOLD_BG   = PatternFill("solid", fgColor="F59E0B")
RED_BG    = PatternFill("solid", fgColor="EF4444")
INPUT_BG  = PatternFill("solid", fgColor="FFF9E6")
LIGHT_BG  = PatternFill("solid", fgColor="F8FAFC")
WHITE_BG  = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="1E293B", size=13)
LABEL_FONT    = Font(name="Calibri", bold=True, color="475569", size=10)
DATA_FONT     = Font(name="Calibri", color="334155", size=10)
FORMULA_FONT  = Font(name="Calibri", color="0F766E", size=10)
BORDER_THIN   = Border(
    bottom=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="E2E8F0"),
)
FMT_USD    = '$#,##0.00'
FMT_USD4   = '$#,##0.0000'
FMT_PCT    = '0.00%'
FMT_PCT4   = '0.0000%'
FMT_SHARES = '#,##0.0000'
FMT_INT    = '#,##0'

MO_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
CONFIG_FILE = os.path.join(SCRIPT_DIR, "spreadsheet_config.txt")


def style_header_row(ws, row, max_col, fill=DARK_BG):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def auto_width(ws, min_width=10, max_width=22):
    for col in ws.columns:
        best = min_width
        for cell in col:
            if cell.value:
                best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[get_column_letter(col[0].column)].width = best


# ============================================================
# DATA LOADING
# ============================================================
def load_data(db, tickers_alloc, years):
    """Query DB and return structured data.

    Returns:
        ticker_info : dict  {symbol: {id, name, alloc}}
        months      : list  of (year, month) sorted chronologically
        prices      : dict  {symbol: {(y,m): {high, close}}}
        dividends   : dict  {symbol: {(y,m): total_div_amount}}
        mm_rates    : dict  {(y,m): rate}   (rate as percentage, e.g. 5.33)
    """
    now = date.today()
    # Stop at prior year-end (December of last complete year)
    end_year, end_month = now.year - 1, 12
    start_year = end_year - years + 1
    start_month = 1

    ticker_info = {}
    for sym, pct in tickers_alloc:
        tk = db.query(Ticker).filter(Ticker.symbol == sym.upper()).first()
        if not tk:
            print(f"  WARNING: ticker '{sym}' not found in DB — skipping")
            continue
        ticker_info[sym.upper()] = {"id": tk.id, "name": tk.name or "", "alloc": pct}

    if not ticker_info:
        print("ERROR: no valid tickers found. Exiting.")
        sys.exit(1)

    # Build month list
    months = []
    y, m = start_year, start_month
    while (y, m) <= (end_year, end_month):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1; y += 1

    # Prices
    prices = {}
    for sym, info in ticker_info.items():
        prices[sym] = {}
        rows = (db.query(MonthlyPrice)
                .filter(MonthlyPrice.ticker_id == info["id"])
                .order_by(MonthlyPrice.year, MonthlyPrice.month)
                .all())
        for r in rows:
            prices[sym][(r.year, r.month)] = {"high": r.high or 0, "close": r.close or 0}

    # Dividends — group by (year, month of pay_date)
    dividends = {}
    for sym, info in ticker_info.items():
        dividends[sym] = defaultdict(float)
        rows = (db.query(Dividend)
                .filter(Dividend.ticker_id == info["id"])
                .order_by(Dividend.pay_date)
                .all())
        for r in rows:
            dividends[sym][(r.pay_date.year, r.pay_date.month)] += r.amount

    # MM rates
    mm_rates = {}
    rows = db.query(MonthlyMoneyMarketRate).order_by(
        MonthlyMoneyMarketRate.year, MonthlyMoneyMarketRate.month).all()
    for r in rows:
        mm_rates[(r.year, r.month)] = r.rate   # stored as percentage (e.g. 5.33)

    return ticker_info, months, prices, dividends, mm_rates


# ============================================================
# SHEET 1: SETUP
# ============================================================
def build_setup(wb, ticker_info, months, prices, dividends, mm_rates,
                monthly_amt, years, tax_rate, annual_growth=0):
    ws = wb.active
    ws.title = "Setup"
    ws.sheet_properties.tabColor = "10B981"

    # --- Parameters ---
    ws["A1"] = "Portfolio Simulator — Test Spreadsheet"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    params = [
        ("Monthly Investment", monthly_amt, FMT_USD),
        ("Investment Years", years, FMT_INT),
        ("Tax Rate", tax_rate / 100, FMT_PCT),
        ("Total Months", len(months), FMT_INT),
        ("Annual Growth", annual_growth, FMT_USD),
    ]
    for i, (label, val, fmt) in enumerate(params):
        r = 3 + i
        ws.cell(r, 1, label).font = LABEL_FONT
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        c.font = DATA_FONT
        c.fill = INPUT_BG

    # --- Ticker allocation table ---
    syms = list(ticker_info.keys())
    tk_row = 9
    ws.cell(tk_row, 1, "Ticker").font = LABEL_FONT
    ws.cell(tk_row, 2, "Name").font = LABEL_FONT
    ws.cell(tk_row, 3, "Allocation %").font = LABEL_FONT
    for i, sym in enumerate(syms):
        r = tk_row + 1 + i
        ws.cell(r, 1, sym).font = DATA_FONT
        ws.cell(r, 2, ticker_info[sym]["name"]).font = DATA_FONT
        c = ws.cell(r, 3, ticker_info[sym]["alloc"] / 100)
        c.number_format = FMT_PCT
        c.font = DATA_FONT
        c.fill = INPUT_BG

    # --- Monthly data table ---
    data_start_row = tk_row + 1 + len(syms) + 2
    # Header row
    hdr = data_start_row
    ws.cell(hdr, 1, "Month").font = LABEL_FONT
    ws.cell(hdr, 2, "MM Rate").font = LABEL_FONT
    col = 3
    for sym in syms:
        ws.cell(hdr, col, f"{sym} High").font = LABEL_FONT
        ws.cell(hdr, col + 1, f"{sym} Close").font = LABEL_FONT
        ws.cell(hdr, col + 2, f"{sym} Div/Share").font = LABEL_FONT
        col += 3
    style_header_row(ws, hdr, col - 1, fill=PatternFill("solid", fgColor="334155"))

    # Data rows
    for mi, (y, m) in enumerate(months):
        r = hdr + 1 + mi
        ws.cell(r, 1, f"{MO_NAMES[m-1]} {y}").font = DATA_FONT
        rate = mm_rates.get((y, m), 0)
        c = ws.cell(r, 2, rate / 100)     # store as decimal for formulas
        c.number_format = FMT_PCT4
        c.font = DATA_FONT
        col = 3
        for sym in syms:
            pd = prices[sym].get((y, m), {"high": 0, "close": 0})
            dv = dividends[sym].get((y, m), 0)
            ws.cell(r, col, pd["high"]).font = DATA_FONT
            ws.cell(r, col, pd["high"]).number_format = FMT_USD
            ws.cell(r, col + 1, pd["close"]).font = DATA_FONT
            ws.cell(r, col + 1, pd["close"]).number_format = FMT_USD
            ws.cell(r, col + 2, dv).font = DATA_FONT
            ws.cell(r, col + 2, dv).number_format = FMT_USD4
            col += 3
        if mi % 2 == 0:
            for c2 in range(1, col):
                ws.cell(r, c2).fill = LIGHT_BG

    ws.freeze_panes = ws.cell(hdr + 1, 1)
    auto_width(ws)

    # Return references for other sheets
    return {
        "amt_cell": "Setup!$B$3",
        "growth_cell": "Setup!$B$7",
        "tax_cell": "Setup!$B$5",
        "alloc_start_row": tk_row + 1,
        "data_hdr_row": hdr,
        "data_start_row": hdr + 1,
        "syms": syms,
        "n_months": len(months),
        "months": months,
        "start_year": months[0][0],
    }


# ============================================================
# SHEET 2: MONTHLY SIMULATION
# ============================================================
def build_monthly_sim(wb, refs):
    ws = wb.create_sheet("Monthly Sim")
    ws.sheet_properties.tabColor = "3B82F6"

    syms = refs["syms"]
    n = refs["n_months"]
    setup_data_start = refs["data_start_row"]

    # Build header
    hdr_row = 1
    headers = ["Month", "MM Rate", "Prior Div Bal", "MM Interest", "Div Bal After Int",
               "Month Budget"]
    for sym in syms:
        headers += [f"{sym} Accum In", f"{sym} Allocated", f"{sym} Accum Bal",
                    f"{sym} Shares", f"{sym} Spent",
                    f"{sym} Accum Out", f"{sym} Cum Shares",
                    f"{sym} Divs", f"{sym} Value"]
    headers += ["Total Shares", "Monthly Divs", "Div Bal End",
                "Cum Invested", "Equity Value", "MM-Only Bal"]

    for ci, h in enumerate(headers, 1):
        ws.cell(hdr_row, ci, h)
    style_header_row(ws, hdr_row, len(headers))

    # Column indices (1-based)
    C_MONTH  = 1
    C_RATE   = 2
    C_PRIOR  = 3
    C_MMINT  = 4
    C_DVAFT  = 5
    C_BUDGET = 6   # Month Budget = flat base amt (no aggregate carryover)
    # Per-ticker block starts at col 7, each ticker takes 9 cols
    def tk_col(ti, offset):
        return 7 + ti * 9 + offset
    # offset: 0=Accum In, 1=Allocated, 2=Accum Bal, 3=Shares, 4=Spent,
    #         5=Accum Out, 6=Cum Shares, 7=Divs, 8=Value
    C_TSHRS  = 7 + len(syms) * 9       # Total Shares (sum of all ticker Cum Shares)
    C_MDIVS  = C_TSHRS + 1
    C_DVEND  = C_TSHRS + 2
    C_CINV   = C_TSHRS + 3
    C_PV     = C_TSHRS + 4
    C_MMONLY = C_TSHRS + 5

    # Reference helpers — Setup sheet data columns
    # Setup col 1=Month, 2=MM Rate, then per ticker 3 cols each starting at 3
    def setup_col(ti, offset):
        """offset: 0=High, 1=Close, 2=Div"""
        return 3 + ti * 3 + offset

    for mi in range(n):
        r = hdr_row + 1 + mi  # data row in this sheet
        sr = setup_data_start + mi  # data row in Setup sheet

        # A: Month label (reference Setup)
        ws.cell(r, C_MONTH, f"=Setup!A{sr}").font = DATA_FONT

        # B: MM Rate (reference Setup col B)
        ws.cell(r, C_RATE, f"=Setup!B{sr}").font = DATA_FONT
        ws.cell(r, C_RATE).number_format = FMT_PCT4

        # C: Prior Div Bal
        if mi == 0:
            ws.cell(r, C_PRIOR, 0).font = DATA_FONT
        else:
            # prior row's Div Bal End
            ws.cell(r, C_PRIOR,
                     f"={get_column_letter(C_DVEND)}{r-1}").font = DATA_FONT
        ws.cell(r, C_PRIOR).number_format = FMT_USD

        # D: MM Interest = ROUND(Prior * Rate / 12, 2)
        rc = get_column_letter(C_RATE)
        pc = get_column_letter(C_PRIOR)
        ws.cell(r, C_MMINT,
                 f"=ROUND({pc}{r}*{rc}{r}/12,2)").font = FORMULA_FONT
        ws.cell(r, C_MMINT).number_format = FMT_USD

        # E: Div Bal After Interest = Prior + MM Interest
        mc = get_column_letter(C_MMINT)
        ws.cell(r, C_DVAFT,
                 f"={pc}{r}+{mc}{r}").font = FORMULA_FONT
        ws.cell(r, C_DVAFT).number_format = FMT_USD

        # F: Month Budget = base + yearOffset * annual growth
        budget_c = get_column_letter(C_BUDGET)
        start_year = refs["start_year"]
        ws.cell(r, C_BUDGET,
                 f"={refs['amt_cell']}+(VALUE(RIGHT(Setup!A{sr},4))-{start_year})*{refs['growth_cell']}").font = FORMULA_FONT
        ws.cell(r, C_BUDGET).number_format = FMT_USD

        # Per-ticker columns (accumulate-then-buy)
        for ti, sym in enumerate(syms):
            c_ain   = tk_col(ti, 0)  # Accum In (prior month's Accum Out)
            c_alloc = tk_col(ti, 1)  # Allocated $ this month
            c_abal  = tk_col(ti, 2)  # Accum Bal = Accum In + Allocated
            c_sh    = tk_col(ti, 3)  # Shares bought (integer)
            c_spent = tk_col(ti, 4)  # Spent = shares × high
            c_aout  = tk_col(ti, 5)  # Accum Out = Accum Bal - Spent
            c_cum   = tk_col(ti, 6)  # Cum shares
            c_div   = tk_col(ti, 7)  # Dividends
            c_val   = tk_col(ti, 8)  # Value

            # Alloc cell in Setup (row tk_row+1+ti, col 3)
            alloc_ref = f"Setup!$C${refs['alloc_start_row'] + ti}"
            # High price in Setup
            high_col = get_column_letter(setup_col(ti, 0))
            close_col = get_column_letter(setup_col(ti, 1))
            div_col = get_column_letter(setup_col(ti, 2))

            # Accum In = prior month's Accum Out (0 for first month)
            ain_c = get_column_letter(c_ain)
            aout_c = get_column_letter(c_aout)
            if mi == 0:
                ws.cell(r, c_ain, 0).font = DATA_FONT
            else:
                ws.cell(r, c_ain,
                         f"={aout_c}{r-1}").font = FORMULA_FONT
            ws.cell(r, c_ain).number_format = FMT_USD

            # Allocated = monthBudget * alloc%
            alloc_c = get_column_letter(c_alloc)
            ws.cell(r, c_alloc,
                     f"={budget_c}{r}*{alloc_ref}").font = FORMULA_FONT
            ws.cell(r, c_alloc).number_format = FMT_USD

            # Accum Bal = ROUND(Accum In + Allocated, 2)
            abal_c = get_column_letter(c_abal)
            ws.cell(r, c_abal,
                     f"=ROUND({ain_c}{r}+{alloc_c}{r},2)").font = FORMULA_FONT
            ws.cell(r, c_abal).number_format = FMT_USD

            # Shares = IF(high>0, INT(accumBal/high), 0) — integer shares only
            sh_c = get_column_letter(c_sh)
            ws.cell(r, c_sh,
                     f"=IF(Setup!{high_col}{sr}>0,INT({abal_c}{r}/Setup!{high_col}{sr}),0)"
                     ).font = FORMULA_FONT
            ws.cell(r, c_sh).number_format = FMT_INT

            # Spent = Shares * High price
            spent_c = get_column_letter(c_spent)
            ws.cell(r, c_spent,
                     f"={sh_c}{r}*Setup!{high_col}{sr}").font = FORMULA_FONT
            ws.cell(r, c_spent).number_format = FMT_USD

            # Accum Out = ROUND(Accum Bal - Spent, 2) — stays in this ticker's bucket
            ws.cell(r, c_aout,
                     f"=ROUND({abal_c}{r}-{spent_c}{r},2)").font = FORMULA_FONT
            ws.cell(r, c_aout).number_format = FMT_USD

            # Cum Shares
            cum_c = get_column_letter(c_cum)
            if mi == 0:
                ws.cell(r, c_cum, f"={sh_c}{r}").font = FORMULA_FONT
            else:
                ws.cell(r, c_cum,
                         f"={cum_c}{r-1}+{sh_c}{r}").font = FORMULA_FONT
            ws.cell(r, c_cum).number_format = FMT_INT

            # Dividends = Cum Shares * Div/Share (from Setup)
            ws.cell(r, c_div,
                     f"={cum_c}{r}*Setup!{div_col}{sr}").font = FORMULA_FONT
            ws.cell(r, c_div).number_format = FMT_USD

            # Value = Cum Shares * Close Price
            ws.cell(r, c_val,
                     f"={cum_c}{r}*Setup!{close_col}{sr}").font = FORMULA_FONT
            ws.cell(r, c_val).number_format = FMT_USD

        # Total Shares = SUM of all per-ticker Cum Shares columns
        cum_refs = "+".join(
            f"{get_column_letter(tk_col(ti,6))}{r}" for ti in range(len(syms)))
        ws.cell(r, C_TSHRS, f"={cum_refs}").font = FORMULA_FONT
        ws.cell(r, C_TSHRS).number_format = FMT_INT

        # Monthly Divs = SUM of all ticker div columns
        div_refs = "+".join(
            f"{get_column_letter(tk_col(ti,7))}{r}" for ti in range(len(syms)))
        ws.cell(r, C_MDIVS, f"={div_refs}").font = FORMULA_FONT
        ws.cell(r, C_MDIVS).number_format = FMT_USD

        # Div Bal End = Div Bal After Int + Monthly Divs
        dvc = get_column_letter(C_DVAFT)
        mdc = get_column_letter(C_MDIVS)
        ws.cell(r, C_DVEND,
                 f"={dvc}{r}+{mdc}{r}").font = FORMULA_FONT
        ws.cell(r, C_DVEND).number_format = FMT_USD

        # Cum Invested = running total of actual $ spent (not allocated)
        spent_refs = "+".join(
            f"{get_column_letter(tk_col(ti,4))}{r}" for ti in range(len(syms)))
        civc = get_column_letter(C_CINV)
        if mi == 0:
            ws.cell(r, C_CINV, f"={spent_refs}").font = FORMULA_FONT
        else:
            ws.cell(r, C_CINV,
                     f"={civc}{r-1}+{spent_refs}").font = FORMULA_FONT
        ws.cell(r, C_CINV).number_format = FMT_USD

        # Equity Value = SUM of all ticker value columns
        val_refs = "+".join(
            f"{get_column_letter(tk_col(ti,8))}{r}" for ti in range(len(syms)))
        ws.cell(r, C_PV, f"={val_refs}").font = FORMULA_FONT
        ws.cell(r, C_PV).number_format = FMT_USD

        # MM-Only Bal = ROUND(prior*(1+rate/12),2) + monthBudget
        # Note: MM benchmark uses same growing budget as equity simulation
        mmc = get_column_letter(C_MMONLY)
        if mi == 0:
            ws.cell(r, C_MMONLY,
                     f"=ROUND(0*(1+{rc}{r}/12),2)+{budget_c}{r}"
                     ).font = FORMULA_FONT
        else:
            ws.cell(r, C_MMONLY,
                     f"=ROUND({mmc}{r-1}*(1+{rc}{r}/12),2)+{budget_c}{r}"
                     ).font = FORMULA_FONT
        ws.cell(r, C_MMONLY).number_format = FMT_USD

        # Alternate row shading
        if mi % 2 == 0:
            for c2 in range(1, C_MMONLY + 1):
                if not ws.cell(r, c2).fill or ws.cell(r, c2).fill == PatternFill():
                    ws.cell(r, c2).fill = LIGHT_BG

    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    auto_width(ws)

    return {
        "hdr_row": hdr_row,
        "data_start_row": hdr_row + 1,
        "C_RATE": C_RATE, "C_MMINT": C_MMINT, "C_MDIVS": C_MDIVS,
        "C_DVEND": C_DVEND, "C_CINV": C_CINV, "C_PV": C_PV, "C_MMONLY": C_MMONLY,
        "C_TSHRS": C_TSHRS, "C_BUDGET": C_BUDGET,
        "tk_col": tk_col,
    }


# ============================================================
# SHEET 3: YEAR SUMMARY
# ============================================================
def build_year_summary(wb, refs, sim_refs, months):
    ws = wb.create_sheet("Year Summary")
    ws.sheet_properties.tabColor = "8B5CF6"

    # Determine unique years
    year_set = sorted(set(y for y, m in months))
    sim_start = sim_refs["data_start_row"]

    headers = ["Year", "Invested", "Cum Invested", "Dividends", "MM Interest",
               "End Stock Value", "End Div Bal"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    style_header_row(ws, 1, len(headers))

    # We need to know which rows in Monthly Sim correspond to each year
    # Build a map: year -> (first_row, last_row) in Monthly Sim
    year_rows = {}
    for mi, (y, m) in enumerate(months):
        r = sim_start + mi
        if y not in year_rows:
            year_rows[y] = [r, r]
        year_rows[y][1] = r

    sim_sheet = "'Monthly Sim'"
    for yi, year in enumerate(year_set):
        r = 2 + yi
        fr, lr = year_rows[year]

        ws.cell(r, 1, year).font = DATA_FONT
        ws.cell(r, 1).number_format = FMT_INT

        # Invested = SUM of all per-ticker Spent columns for this year's rows
        # (Spent = shares × high, offset 4 in per-ticker block)
        inv_parts = []
        for ti in range(len(refs["syms"])):
            col_letter = get_column_letter(sim_refs["tk_col"](ti, 4))
            inv_parts.append(f"SUM({sim_sheet}!{col_letter}{fr}:{col_letter}{lr})")
        ws.cell(r, 2, "=" + "+".join(inv_parts)).font = FORMULA_FONT
        ws.cell(r, 2).number_format = FMT_USD

        # Cum Invested = reference the last month's Cum Invested
        cinv_c = get_column_letter(sim_refs["C_CINV"])
        ws.cell(r, 3, f"={sim_sheet}!{cinv_c}{lr}").font = FORMULA_FONT
        ws.cell(r, 3).number_format = FMT_USD

        # Dividends = SUM of Monthly Divs column
        mdivs_c = get_column_letter(sim_refs["C_MDIVS"])
        ws.cell(r, 4,
                 f"=SUM({sim_sheet}!{mdivs_c}{fr}:{mdivs_c}{lr})").font = FORMULA_FONT
        ws.cell(r, 4).number_format = FMT_USD

        # MM Interest = SUM of MM Interest column
        mmint_c = get_column_letter(sim_refs["C_MMINT"])
        ws.cell(r, 5,
                 f"=SUM({sim_sheet}!{mmint_c}{fr}:{mmint_c}{lr})").font = FORMULA_FONT
        ws.cell(r, 5).number_format = FMT_USD

        # End Stock Value = last month's Portfolio Value
        pv_c = get_column_letter(sim_refs["C_PV"])
        ws.cell(r, 6, f"={sim_sheet}!{pv_c}{lr}").font = FORMULA_FONT
        ws.cell(r, 6).number_format = FMT_USD

        # End Div Bal = last month's Div Bal End
        dvend_c = get_column_letter(sim_refs["C_DVEND"])
        ws.cell(r, 7, f"={sim_sheet}!{dvend_c}{lr}").font = FORMULA_FONT
        ws.cell(r, 7).number_format = FMT_USD

        if yi % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = LIGHT_BG

    ws.freeze_panes = ws.cell(2, 1)
    auto_width(ws)

    return {"data_start_row": 2, "n_years": len(year_set), "years": year_set}


# ============================================================
# SHEET 4: TAX IMPACT
# ============================================================
def build_tax_impact(wb, refs, yr_refs):
    ws = wb.create_sheet("Tax Impact")
    ws.sheet_properties.tabColor = "EF4444"

    headers = ["Year", "Dividends", "Tax on Dividends", "MM Interest",
               "Tax on Interest", "Total Taxes"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    style_header_row(ws, 1, len(headers), fill=PatternFill("solid", fgColor="7F1D1D"))

    tax_ref = refs["tax_cell"]
    yr_start = yr_refs["data_start_row"]
    ys_sheet = "'Year Summary'"

    for yi in range(yr_refs["n_years"]):
        r = 2 + yi
        yr_r = yr_start + yi

        # Year
        ws.cell(r, 1, f"={ys_sheet}!A{yr_r}").font = DATA_FONT
        ws.cell(r, 1).number_format = FMT_INT

        # Dividends (from Year Summary col D)
        ws.cell(r, 2, f"={ys_sheet}!D{yr_r}").font = FORMULA_FONT
        ws.cell(r, 2).number_format = FMT_USD

        # Tax on Dividends
        ws.cell(r, 3, f"=B{r}*{tax_ref}").font = FORMULA_FONT
        ws.cell(r, 3).number_format = FMT_USD

        # MM Interest (from Year Summary col E)
        ws.cell(r, 4, f"={ys_sheet}!E{yr_r}").font = FORMULA_FONT
        ws.cell(r, 4).number_format = FMT_USD

        # Tax on Interest
        ws.cell(r, 5, f"=D{r}*{tax_ref}").font = FORMULA_FONT
        ws.cell(r, 5).number_format = FMT_USD

        # Total Taxes
        ws.cell(r, 6, f"=C{r}+E{r}").font = FORMULA_FONT
        ws.cell(r, 6).number_format = FMT_USD

        if yi % 2 == 0:
            for c in range(1, 7):
                ws.cell(r, c).fill = LIGHT_BG

    # Totals row
    tr = 2 + yr_refs["n_years"]
    ws.cell(tr, 1, "TOTAL").font = Font(name="Calibri", bold=True, size=11)
    for c in range(2, 7):
        cl = get_column_letter(c)
        ws.cell(tr, c, f"=SUM({cl}2:{cl}{tr-1})")
        ws.cell(tr, c).font = Font(name="Calibri", bold=True, color="EF4444", size=11)
        ws.cell(tr, c).number_format = FMT_USD
    for c in range(1, 7):
        ws.cell(tr, c).border = Border(top=Side(style="medium", color="1E293B"))

    ws.freeze_panes = ws.cell(2, 1)
    auto_width(ws)

    return {"data_start_row": 2}


# ============================================================
# SHEET 5: ANNUAL RETURNS
# ============================================================
def build_annual_returns(wb, refs, yr_refs):
    ws = wb.create_sheet("Annual Returns")
    ws.sheet_properties.tabColor = "F59E0B"

    ys_sheet = "'Year Summary'"
    yr_start = yr_refs["data_start_row"]

    # For each year we write a summary row + a calculation breakdown block
    r = 1

    # Main header (After-Tax Value and After-Tax Return removed)
    main_headers = ["Year", "Invested", "Dividends", "MM Interest", "Stock Value",
                    "Portfolio Value", "Pre-Tax Return"]
    for ci, h in enumerate(main_headers, 1):
        ws.cell(r, ci, h)
    style_header_row(ws, r, len(main_headers))
    r += 1

    for yi in range(yr_refs["n_years"]):
        yr_r = yr_start + yi    # row in Year Summary

        # --- Summary row ---
        sr = r  # remember this row for the detail formulas

        # Year
        ws.cell(r, 1, f"={ys_sheet}!A{yr_r}").font = Font(bold=True, size=11)
        ws.cell(r, 1).number_format = FMT_INT
        # Invested
        ws.cell(r, 2, f"={ys_sheet}!B{yr_r}").font = FORMULA_FONT
        ws.cell(r, 2).number_format = FMT_USD
        # Dividends
        ws.cell(r, 3, f"={ys_sheet}!D{yr_r}").font = FORMULA_FONT
        ws.cell(r, 3).number_format = FMT_USD
        # MM Interest
        ws.cell(r, 4, f"={ys_sheet}!E{yr_r}").font = FORMULA_FONT
        ws.cell(r, 4).number_format = FMT_USD
        # Stock Value (end of year)
        ws.cell(r, 5, f"={ys_sheet}!F{yr_r}").font = FORMULA_FONT
        ws.cell(r, 5).number_format = FMT_USD
        # Portfolio Value = Stock Value + Div Bal
        ws.cell(r, 6, f"={ys_sheet}!F{yr_r}+{ys_sheet}!G{yr_r}").font = FORMULA_FONT
        ws.cell(r, 6).number_format = FMT_USD

        # Pre-Tax Return — will reference detail rows below
        r += 1

        # --- Calculation detail block ---
        detail_start = r
        details = [
            ("Beginning Stock Value", None),
            ("Stock Gain", "End Stock Value - Beginning - Invested"),
            ("Avg Invested Capital", "Invested x 0.542 (DCA mid-year)"),
            ("Base Capital (Denominator)", "Beginning Stock Value + Avg Invested"),
            ("Total Gain (Numerator)", "Stock Gain + Dividends + MM Interest"),
            ("Pre-Tax Return", "Total Gain / Base Capital"),
        ]

        for di, (label, note) in enumerate(details):
            dr = detail_start + di
            ws.cell(dr, 1, "").font = DATA_FONT   # indent
            ws.cell(dr, 2, label).font = Font(name="Calibri", italic=True,
                                               color="64748B", size=10)
            if note:
                ws.cell(dr, 4, note).font = Font(name="Calibri", italic=True,
                                                   color="94A3B8", size=9)
            for c in range(1, len(main_headers) + 1):
                ws.cell(dr, c).fill = PatternFill("solid", fgColor="F1F5F9")

        # Row references for detail formulas
        r_beg    = detail_start + 0
        r_sg     = detail_start + 1
        r_avg    = detail_start + 2
        r_base   = detail_start + 3
        r_tgain  = detail_start + 4
        r_ptret  = detail_start + 5

        # Beginning Stock Value = prior year's End Stock Value (or 0 for first year)
        if yi == 0:
            ws.cell(r_beg, 3, 0).font = FORMULA_FONT
        else:
            prev_yr_r = yr_start + yi - 1
            ws.cell(r_beg, 3, f"={ys_sheet}!F{prev_yr_r}").font = FORMULA_FONT
        ws.cell(r_beg, 3).number_format = FMT_USD

        # Stock Gain = End Stock Value - Beginning - Invested
        ws.cell(r_sg, 3,
                 f"=E{sr}-C{r_beg}-B{sr}").font = FORMULA_FONT
        ws.cell(r_sg, 3).number_format = FMT_USD

        # Avg Invested Capital = Invested * 0.542
        ws.cell(r_avg, 3, f"=B{sr}*0.542").font = FORMULA_FONT
        ws.cell(r_avg, 3).number_format = FMT_USD

        # Base Capital = Beginning + Avg Invested
        ws.cell(r_base, 3, f"=C{r_beg}+C{r_avg}").font = FORMULA_FONT
        ws.cell(r_base, 3).number_format = FMT_USD

        # Total Gain = Stock Gain + Dividends + MM Interest
        ws.cell(r_tgain, 3,
                 f"=C{r_sg}+C{sr}+D{sr}").font = FORMULA_FONT
        ws.cell(r_tgain, 3).number_format = FMT_USD

        # Pre-Tax Return = Total Gain / Base Capital
        ws.cell(r_ptret, 3,
                 f"=IF(C{r_base}>0,C{r_tgain}/C{r_base},0)").font = FORMULA_FONT
        ws.cell(r_ptret, 3).number_format = FMT_PCT

        # Fill in the summary row's Pre-Tax Return column (G)
        ws.cell(sr, 7, f"=C{r_ptret}").font = Font(bold=True, color="10B981", size=11)
        ws.cell(sr, 7).number_format = FMT_PCT

        # Highlight the summary row
        for c in range(1, len(main_headers) + 1):
            ws.cell(sr, c).fill = PatternFill("solid", fgColor="ECFDF5")
            ws.cell(sr, c).border = Border(
                bottom=Side(style="thin", color="10B981"))

        r = detail_start + len(details) + 1  # blank row between years

    ws.freeze_panes = ws.cell(2, 1)
    auto_width(ws, min_width=14)


# ============================================================
# MAIN
# ============================================================
def read_config():
    """Read parameters from spreadsheet_config.txt if it exists.

    Returns (tickers_alloc, monthly_amt, years, tax_rate) or None if
    the config file is missing or has errors.
    """
    if not os.path.exists(CONFIG_FILE):
        return None

    config = {}
    with open(CONFIG_FILE) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            config[key.strip().lower()] = value.strip()

    required = ["tickers", "monthly_amount", "years", "tax_rate"]
    missing = [k for k in required if k not in config]
    if missing:
        print(f"  Config file missing fields: {', '.join(missing)}")
        print(f"  Falling back to interactive prompts...")
        return None

    # Parse tickers
    tickers_alloc = []
    for part in config["tickers"].split(","):
        part = part.strip()
        if ":" not in part:
            print(f"  Invalid ticker format '{part}' in config — use TICKER:PCT")
            return None
        sym, pct = part.split(":", 1)
        tickers_alloc.append((sym.strip().upper(), float(pct.strip())))

    monthly_amt = float(config["monthly_amount"])
    years = int(config["years"])
    tax_rate = float(config["tax_rate"])
    annual_growth = float(config.get("annual_growth", 0))

    return tickers_alloc, monthly_amt, years, tax_rate, annual_growth


def main():
    print("=" * 60)
    print("  Portfolio Simulator — Test Spreadsheet Generator")
    print("=" * 60)
    print()

    # --- Try config file first, then interactive prompts ---
    cfg = read_config()
    if cfg:
        tickers_alloc, monthly_amt, years, tax_rate, annual_growth = cfg
        print(f"  Read from: {CONFIG_FILE}")
    else:
        # Interactive prompts (fallback)
        raw = input("Enter tickers and allocation % (e.g. XLK:60,XLV:40): ").strip()
        if not raw:
            print("No input. Exiting.")
            return
        tickers_alloc = []
        for part in raw.split(","):
            part = part.strip()
            if ":" not in part:
                print(f"  Invalid format '{part}' — use TICKER:PCT")
                return
            sym, pct = part.split(":", 1)
            tickers_alloc.append((sym.strip().upper(), float(pct.strip())))

        amt_str = input("Monthly investment amount [$1000]: ").strip()
        monthly_amt = float(amt_str) if amt_str else 1000.0

        yr_str = input("Investment period in years [10]: ").strip()
        years = int(yr_str) if yr_str else 10

        tax_str = input("Tax rate % [30]: ").strip()
        tax_rate = float(tax_str) if tax_str else 30.0

        growth_str = input("Annual deposit growth [$0]: ").strip()
        annual_growth = float(growth_str) if growth_str else 0.0

    total_pct = sum(p for _, p in tickers_alloc)
    if abs(total_pct - 100) > 0.01:
        print(f"  WARNING: allocations sum to {total_pct}%, not 100%")

    print()
    print(f"  Tickers: {', '.join(f'{s} {p}%' for s,p in tickers_alloc)}")
    print(f"  Amount:  ${monthly_amt:,.0f}/month for {years} years")
    print(f"  Growth:  +${annual_growth:,.0f}/year")
    print(f"  Tax:     {tax_rate}%")
    print()

    # --- Load data from DB ---
    print("Connecting to database...")
    db = SessionLocal()
    try:
        ticker_info, months, prices, dividends, mm_rates = load_data(
            db, tickers_alloc, years)
    finally:
        db.close()

    print(f"  Loaded {len(months)} months of data for "
          f"{', '.join(ticker_info.keys())}")
    print()

    # --- Build workbook ---
    print("Building spreadsheet...")
    wb = Workbook()

    refs = build_setup(wb, ticker_info, months, prices, dividends, mm_rates,
                       monthly_amt, years, tax_rate, annual_growth)
    sim_refs = build_monthly_sim(wb, refs)
    yr_refs = build_year_summary(wb, refs, sim_refs, months)
    tax_refs = build_tax_impact(wb, refs, yr_refs)
    build_annual_returns(wb, refs, yr_refs)

    # --- Save ---
    # If the default file is locked (open in Excel), add a suffix
    out_path = os.path.join(SCRIPT_DIR, "Portfolio_Simulator_Test.xlsx")
    if os.path.exists(out_path):
        try:
            with open(out_path, "a"):
                pass
        except PermissionError:
            out_path = os.path.join(SCRIPT_DIR, "Portfolio_Simulator_Test_new.xlsx")
            print(f"  (Previous file is open in Excel, saving as new file)")
    wb.save(out_path)
    print(f"  Saved to: {out_path}")
    print()
    print("Open in Excel and verify formulas. Change tax rate or")
    print("monthly amount on the Setup sheet — all sheets update.")


if __name__ == "__main__":
    main()
